"""Nó 1 — extração e detecção de tipo, a partir dos documentos enviados.

Aceita PDF, Excel e imagem "como estão" (fricção mínima). Texto sai de PDF (pdfplumber) e
Excel (openpyxl); imagem é aceita e guardada, mas sem OCR configurado (extração best-effort).

Detecção do tipo de documento de transporte é AUTOMÁTICA (o próprio documento diz o que é —
B/L "Bill of Lading", AWB "Air Waybill", CRT rodoviário) — sem perguntar o modal antes.

A extração de campos aqui é heurística/determinística (regex). O Nó 1 "pleno" com LLM
(extração estruturada de qualidade) é um upgrade plugável — quando ANTHROPIC_API_KEY existir,
dá para trocar `extrair_campos` por uma chamada ao modelo. O que a demo precisa provar (a
citação normativa ao lado do apontamento) já roda sobre o NCM, que é regex robusto.
"""
from __future__ import annotations

import io
import re

import openpyxl
import pdfplumber
import xlrd


def abas_texto(nome: str, mime: str, conteudo: bytes) -> dict[str, str]:
    """Texto por aba/planilha — chave = nome da aba. Suporta .xlsx (openpyxl) e .xls (xlrd).
    Documentos reais de comex vêm com Invoice e Packing List em ABAS separadas do mesmo arquivo.
    Para PDF/imagem retorna {'documento': <texto extraído>} (uma "aba" única)."""
    nome_l = (nome or "").lower()
    try:
        if nome_l.endswith(".xls") and not nome_l.endswith(".xlsx"):
            wb = xlrd.open_workbook(file_contents=conteudo)
            out = {}
            for sh in wb.sheets():
                linhas = [" ".join(str(sh.cell_value(r, c)) for c in range(sh.ncols))
                          for r in range(sh.nrows)]
                out[sh.name] = "\n".join(linhas)
            return out
        if nome_l.endswith(".xlsx") or "spreadsheet" in mime:
            wb = openpyxl.load_workbook(io.BytesIO(conteudo), read_only=True, data_only=True)
            out = {}
            for ws in wb.worksheets:
                out[ws.title] = "\n".join(
                    " ".join(str(c) for c in row if c is not None) for row in ws.iter_rows(values_only=True))
            wb.close()
            return out
    except Exception as e:
        return {"documento": f"[falha ao ler planilha: {e}]"}
    return {"documento": extrair_texto(nome, mime, conteudo)}

# NCM: 8 dígitos, normalmente formatados 0000.00.00 (aceita com/sem pontos).
_RE_NCM = re.compile(r"\b(\d{4}\.?\d{2}\.?\d{2})\b")


def extrair_texto(nome: str, mime: str, conteudo: bytes) -> str:
    nome_l = (nome or "").lower()
    try:
        if nome_l.endswith(".pdf") or "pdf" in mime:
            with pdfplumber.open(io.BytesIO(conteudo)) as pdf:
                return "\n".join((p.extract_text() or "") for p in pdf.pages[:8])
        if nome_l.endswith(".xls") and not nome_l.endswith(".xlsx"):
            wb = xlrd.open_workbook(file_contents=conteudo)
            return "\n".join(" ".join(str(sh.cell_value(r, c)) for c in range(sh.ncols))
                             for sh in wb.sheets() for r in range(sh.nrows))
        if nome_l.endswith(".xlsx") or "spreadsheet" in mime or "excel" in mime:
            wb = openpyxl.load_workbook(io.BytesIO(conteudo), read_only=True, data_only=True)
            partes = []
            for ws in wb.worksheets:
                for row in ws.iter_rows(values_only=True):
                    partes.append(" ".join(str(c) for c in row if c is not None))
            wb.close()
            return "\n".join(partes)
    except Exception as e:  # documento corrompido/ilegível — não quebra o fluxo
        return f"[falha ao extrair texto: {e}]"
    return ""  # imagem ou formato sem extrator de texto (OCR não configurado)


def detectar_tipo_transporte(texto: str) -> str | None:
    """Detecta B/L, AWB ou CRT pelo próprio conteúdo. None = indetectado (humano confirma)."""
    t = (texto or "").lower()
    if any(k in t for k in ("air waybill", "airwaybill", "awb", "conhecimento aéreo", "conhecimento aereo")):
        return "AWB"
    if any(k in t for k in ("bill of lading", "b/l", "conhecimento de embarque", "ocean bill",
                            "vessel", "porto de embarque", "port of loading")):
        return "B/L"
    if any(k in t for k in ("crt", "conhecimento rodoviário", "conhecimento rodoviario",
                            "transporte rodoviário internacional")):
        return "CRT"
    return None


def extrair_ncms(texto: str) -> list[str]:
    """NCMs distintos encontrados, normalizados para 0000.00.00 (regex robusto)."""
    vistos: list[str] = []
    for m in _RE_NCM.findall(texto or ""):
        digitos = re.sub(r"\D", "", m)
        if len(digitos) == 8:
            fmt = f"{digitos[:4]}.{digitos[4:6]}.{digitos[6:]}"
            if fmt not in vistos:
                vistos.append(fmt)
    return vistos


def contexto_ncm(texto: str, ncm_fmt: str) -> str:
    """Descrição da mercadoria (limpa) na linha que menciona o NCM — melhor base para a query
    normativa do que a descrição terse do NCM (muitas vezes 'Outros'). Remove ruído que causa
    deriva semântica (o próprio código NCM, 'Item N:', rótulos), deixando só o produto."""
    digitos = re.sub(r"\D", "", ncm_fmt)
    for linha in (texto or "").splitlines():
        if not (digitos and digitos in re.sub(r"\D", "", linha)):
            continue
        s = re.sub(r"(?i)\bNCM\b.*", "", linha)          # remove "NCM ..." até o fim da linha
        s = re.sub(r"(?i)^\s*item\s*\d+\s*[:.\-]?", "", s)  # remove "Item N:"
        s = re.sub(r"\d{4}\.?\d{2}\.?\d{2}", "", s)        # remove códigos residuais
        s = re.sub(r"\s+", " ", s).strip(" -–:.")
        return s[:200]
    return ""


def extrair_campos(texto: str) -> dict:
    """Campos de conciliação (best-effort, heurístico). Chaves ausentes = não encontrado."""
    campos: dict[str, str] = {}
    m_inv = re.search(r"(?:invoice|fatura)\s*(?:n[oº.:]*|number|no\.?)\s*[:#]?\s*([A-Z0-9\-/]{3,})", texto or "", re.I)
    if m_inv:
        campos["numero"] = m_inv.group(1).strip()
    m_peso = re.search(r"(?:gross\s*weight|peso\s*bruto)\s*[:.]?\s*([\d.,]+)\s*(kg|kgs|k)", texto or "", re.I)
    if m_peso:
        campos["peso_bruto"] = m_peso.group(1).strip()
    m_vol = re.search(r"(?:packages|volumes|pkgs|bultos)\s*[:.]?\s*(\d+)", texto or "", re.I)
    if m_vol:
        campos["volumes"] = m_vol.group(1).strip()
    m_tot = re.search(r"(?:total|amount|valor\s*total)\s*[:.]?\s*(?:USD|US\$|R\$|\$)?\s*([\d.,]{2,})", texto or "", re.I)
    if m_tot:
        campos["valor_total"] = m_tot.group(1).strip()
    # Incoterm (2020) + termo geográfico opcional: 'FOB Ningbo', 'CIF Santos'.
    m_inc = re.search(r"\b(EXW|FCA|FAS|FOB|CFR|CIF|CPT|CIP|DAP|DPU|DDP)\b(?:\s+([A-Za-zÀ-ÿ][A-Za-zÀ-ÿ .'-]{1,28}))?",
                      texto or "", re.I)
    if m_inc:
        local = (m_inc.group(2) or "").strip(" .,-")
        campos["incoterm"] = (m_inc.group(1).upper() + (f" {local}" if local else "")).strip()
    m_frete = re.search(r"freight\s*(prepaid|collect)|frete\s*(pago|a\s*pagar|a\s*cobrar)", texto or "", re.I)
    if m_frete:
        campos["condicao_frete"] = m_frete.group(0).strip()
    m_origem = re.search(
        r"(?:country\s*of\s*origin|made\s*in|pa[ií]s\s*de\s*origem)\s*[:.]?\s*([A-Za-zÀ-ÿ][A-Za-zÀ-ÿ .'-]{2,40})",
        texto or "", re.I)
    if m_origem:
        campos["pais_origem"] = m_origem.group(1).strip(" .,-")
    return campos
