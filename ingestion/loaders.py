"""Loaders plugáveis por tipo de fonte.

Cada loader recebe uma FonteConfig e devolve uma lista de UnidadeNormativa já
chunkada (uma unidade citável por item). Adicionar suporte a um novo tipo de fonte
= registrar um novo loader, sem tocar no pipeline.

Princípio de grounding: NUNCA ingerir conteúdo não verificado. O loader 'file' lê
texto normativo já coletado e conferido (JSON estruturado sob seeds/). O loader
'http' é stub proposital para fontes cuja coleta+parse ainda não foi implementada.
O loader 'ncm_json' consome o JSON oficial da nomenclatura NCM do Portal Único.
"""
from __future__ import annotations

import html as html_mod
import io
import json
import pathlib
import re
import time
from typing import Callable

import httpx

from models import FonteConfig, UnidadeNormativa

SEEDS_DIR = pathlib.Path(__file__).resolve().parent / "seeds"

# Endpoint oficial do JSON da nomenclatura NCM (Portal Único Siscomex).
NCM_JSON_URL = "https://portalunico.siscomex.gov.br/classif/api/publico/nomenclatura/download/json"

# User-Agent de navegador — gov.br às vezes recusa clientes sem UA.
_UA = {"User-Agent": "Mozilla/5.0 (compatible; daleship-compliance-engine/1.0)"}

_LOADERS: dict[str, Callable[[FonteConfig], list[UnidadeNormativa]]] = {}


def register(nome: str):
    def deco(fn: Callable[[FonteConfig], list[UnidadeNormativa]]):
        _LOADERS[nome] = fn
        return fn

    return deco


def get_loader(nome: str) -> Callable[[FonteConfig], list[UnidadeNormativa]]:
    if nome not in _LOADERS:
        raise KeyError(f"Loader '{nome}' não registrado. Disponíveis: {sorted(_LOADERS)}")
    return _LOADERS[nome]


@register("file")
def file_loader(fonte: FonteConfig) -> list[UnidadeNormativa]:
    """Lê unidades normativas verificadas de um JSON sob seeds/.

    Formato esperado: lista de objetos {"identificador": str, "texto": str}.
    """
    if not fonte.caminho:
        raise ValueError(f"Fonte {fonte.orgao}/{fonte.tipo_documento}: loader 'file' exige 'caminho'.")
    caminho = SEEDS_DIR / fonte.caminho
    dados = json.loads(caminho.read_text(encoding="utf-8"))
    return [UnidadeNormativa(identificador=u["identificador"], texto=u["texto"]) for u in dados]


@register("http")
def http_loader(fonte: FonteConfig) -> list[UnidadeNormativa]:
    """Stub: coleta+chunk de fonte HTTP oficial ainda não implementada para esta fonte.

    Exige etapa de verificação humana do conteúdo parseado antes de virar norma
    citável — não ingerir automaticamente sem isso."""
    raise NotImplementedError(
        f"Loader 'http' ainda não implementado para {fonte.fonte_url}. "
        "Colete e verifique o conteúdo, salve como JSON em seeds/ e use o loader 'file'."
    )


# --- NCM (Portal Único) ---

def _desc_hierarquica(digitos: str, mapa: dict[str, str]) -> str:
    """Descrição CONCATENADA da NCM: junta as descrições dos níveis ancestrais (capítulo 2,
    posição 4, subposição 6, item 8) — igual à versão 'concatenada' oficial. Sem isso, a folha
    é só 'Outros' e a busca/classificação não funciona (ex.: cobertor caía em NCM de papel)."""
    partes: list[str] = []
    for n in (2, 4, 6, 8, 10):
        pref = digitos[:n]
        if len(pref) < n:
            break
        d = mapa.get(pref)
        if d and (not partes or d != partes[-1]):
            partes.append(d)
    return " > ".join(partes)


def parse_ncm_payload(payload: dict) -> list[UnidadeNormativa]:
    """Transforma o payload JSON da nomenclatura NCM em unidades citáveis, com a descrição
    HIERÁRQUICA concatenada (não só a folha 'Outros').

    Uma unidade por código. `identificador` = 'NCM <codigo>'. Estrutura:
    {"Nomenclaturas": [{"Codigo","Descricao",...}]} (todos os níveis 2/4/6/8 díg presentes)."""
    itens = payload.get("Nomenclaturas") or payload.get("nomenclaturas") or []
    # 1ª passada: mapa digitos->descrição (sem tags) de TODOS os níveis
    mapa: dict[str, str] = {}
    for item in itens:
        cod = re.sub(r"\D", "", item.get("Codigo") or item.get("codigo") or "")
        d = re.sub(r"<[^>]+>", "", (item.get("Descricao") or item.get("descricao") or "")).strip()
        if cod:
            mapa[cod] = d
    # 2ª passada: unidades com descrição hierárquica
    unidades: list[UnidadeNormativa] = []
    for item in itens:
        codigo = (item.get("Codigo") or item.get("codigo") or "").strip()
        digitos = re.sub(r"\D", "", codigo)
        if not digitos:
            continue
        texto = f"{codigo} — {_desc_hierarquica(digitos, mapa)}"
        unidades.append(UnidadeNormativa(identificador=f"NCM {codigo}", texto=texto))
    return unidades


def _portal_em_parada(resp: httpx.Response) -> bool:
    """Detecta a parada programada diária do Portal Único (01:00–03:00): a chamada
    acaba em parada-programada.json. Não indexar o conteúdo de status.

    (Segue redirects normais, ex.: 307 que só acrescenta ?perfil=PUBLICO — a parada é
    identificada pela URL FINAL, não por qualquer redirect.)"""
    return "parada-programada" in str(resp.url)


@register("ncm_json")
def ncm_json_loader(fonte: FonteConfig) -> list[UnidadeNormativa]:
    """Baixa o JSON oficial da nomenclatura NCM e chunka por código.

    Health-check embutido: se o Portal Único estiver na parada programada, aborta
    com mensagem clara em vez de ingerir o payload de status."""
    with httpx.Client(timeout=180, follow_redirects=True, headers=_UA) as client:
        resp = client.get(NCM_JSON_URL)
        if _portal_em_parada(resp):
            raise RuntimeError(
                "Portal Único em parada programada (01:00–03:00). "
                "Rodar a coleta da NCM fora dessa janela."
            )
        resp.raise_for_status()
        payload = resp.json()
    return parse_ncm_payload(payload)


# --- Tratamento Administrativo na Importação (compilado de anuentes, gov.br/siscomex) ---

# Rótulo por coluna do compilado_ta_anuente (validado contra o cabeçalho no parse).
_TA_COLUNAS = {
    1: "Escopo (produto/condições/modelo LPCO)",
    2: "Fundamentação legal para atuação na importação",
    3: "Tipo de controle administrativo",
    4: "Tipo de LPCO",
    5: "LPCO demanda catálogo",
    6: "Validade do LPCO",
    7: "Tipo de CNPJ no LPCO",
    8: "LPCO retificável",
    9: "LPCO com taxa em outro sistema",
    10: "LPCO com taxa integrada ao PCCE",
    11: "LPCO prévio ao embarque",
    12: "Conferência/inspeção do anuente na DUIMP",
}


def resolver_xlsx_por_marcador(html: str, base_url: str, marcador: str) -> str:
    """Acha, na página/pasta oficial, o link .xlsx cujo href contém `marcador`
    (ex.: 'compilado_ta_anuente') — não assume o nome exato (o arquivo tem o ano)."""
    hrefs = re.findall(r'href="([^"]+\.xlsx)"', html, re.I)
    for href in hrefs:
        if marcador.lower() in href.lower():
            return str(httpx.URL(base_url).join(href))
    raise RuntimeError(f"Nenhum .xlsx contendo '{marcador}' encontrado na página oficial.")


def parse_ta_rows(rows: list, filtro_orgaos: list[str] | None = None) -> list[UnidadeNormativa]:
    """Converte as linhas do compilado de Tratamento Administrativo em unidades citáveis
    (uma por linha = um tratamento por órgão/escopo). Valida o cabeçalho antes: se a
    estrutura da planilha mudar, aborta em vez de mis-parsear (grounding)."""
    if len(rows) < 4:
        raise RuntimeError("Compilado TA com poucas linhas — estrutura inesperada.")
    cab = [(str(c).strip().upper() if c else "") for c in rows[1]]
    if cab[0] != "ÓRGÃO" or "FUNDAMENTAÇÃO LEGAL" not in cab[2]:
        raise RuntimeError("Cabeçalho do compilado TA mudou — abortar para não mis-parsear.")

    filtro = {o.upper() for o in filtro_orgaos} if filtro_orgaos else None
    vistos: dict[str, int] = {}
    unidades: list[UnidadeNormativa] = []
    for r in rows[3:]:
        orgao = (str(r[0]).strip() if r[0] is not None else "")
        escopo = (str(r[1]).strip() if len(r) > 1 and r[1] is not None else "")
        if not orgao or not escopo:
            continue
        if filtro and orgao.upper() not in filtro:
            continue

        partes = [f"Órgão anuente: {orgao}."]
        for idx, rotulo in _TA_COLUNAS.items():
            val = str(r[idx]).strip() if idx < len(r) and r[idx] is not None else ""
            if val:
                partes.append(f"{rotulo}: {re.sub(r'\\s+', ' ', val)}.")
        texto = " ".join(partes)

        base_id = f"Tratamento Administrativo Importação — {orgao}: {re.sub(r'\\s+', ' ', escopo)[:80]}"
        n = vistos.get(base_id, 0) + 1
        vistos[base_id] = n
        ident = base_id if n == 1 else f"{base_id} ({n})"
        unidades.append(UnidadeNormativa(identificador=ident, texto=texto))
    return unidades


@register("tratamento_adm_ta")
def tratamento_adm_ta_loader(fonte: FonteConfig) -> list[UnidadeNormativa]:
    """Coleta o compilado oficial de Tratamento Administrativo de Importação (anuentes).

    fonte_url = a PASTA /informacoes/ que lista o xlsx (fetchável por httpx; a página
    /servicos/ não é). Resolve o link 'compilado_ta_anuente*.xlsx', baixa e parseia.

    params.filtro_orgaos: lista opcional para restringir órgãos (ex.: ['ANVISA','MAPA'])."""
    params = fonte.params or {}
    filtro = params.get("filtro_orgaos")
    with httpx.Client(timeout=180, follow_redirects=True, headers=_UA) as client:
        pagina = client.get(fonte.fonte_url)
        pagina.raise_for_status()
        xlsx_url = resolver_xlsx_por_marcador(pagina.text, str(pagina.url), "compilado_ta_anuente")
        resp = client.get(xlsx_url)
        resp.raise_for_status()
        conteudo = resp.content

    import openpyxl

    wb = openpyxl.load_workbook(io.BytesIO(conteudo), read_only=True, data_only=True)
    aba = "Planilha1" if "Planilha1" in wb.sheetnames else wb.sheetnames[0]
    rows = list(wb[aba].iter_rows(values_only=True))
    wb.close()
    return parse_ta_rows(rows, filtro)


# --- RGI via NESH (Notas Explicativas do Sistema Harmonizado, Receita Federal) ---

def resolver_pdf_nesh(html: str, base_url: str) -> str:
    """Acha, no HTML da página oficial, o link de download do PDF vigente da NESH.

    Não assume o nome do arquivo (a versão pode mudar). Prefere o link cujo texto de
    âncora referencia a IN vigente (2.169/2023); descarta explicitamente a 'versão anterior'.
    """
    ancoras = re.findall(r'<a[^>]*href="([^"]+\.pdf)"[^>]*>(.*?)</a>', html, re.I | re.S)
    candidatos = []
    for href, texto in ancoras:
        texto_limpo = re.sub(r"<[^>]+>", " ", texto)
        if "anterior" in texto_limpo.lower():
            continue
        candidatos.append((href, texto_limpo))
    if not candidatos:
        raise RuntimeError("Nenhum link de PDF da NESH encontrado na página oficial.")

    # Preferência: âncora que cita a IN vigente (2.169 / 2169).
    for href, texto in candidatos:
        if "2.169" in texto or "2169" in texto:
            return httpx.URL(base_url).join(href).__str__()
    # Fallback: primeiro candidato que não é a versão anterior.
    href = candidatos[0][0]
    return httpx.URL(base_url).join(href).__str__()


def parse_rgi_texto(texto: str) -> list[UnidadeNormativa]:
    """Extrai as 6 Regras Gerais de Interpretação do texto da seção RGI da NESH.

    Cada unidade = o enunciado normativo da REGRA N (de 'REGRA N' até 'NOTA EXPLICATIVA').
    Se qualquer uma das 6 faltar, aborta — nunca indexa RGI parcial (grounding)."""
    unidades: list[UnidadeNormativa] = []
    for n in range(1, 7):
        m = re.search(rf"(?m)^REGRA {n}\s*$", texto)
        if not m:
            raise RuntimeError(f"RGI Regra {n} não localizada no texto extraído da NESH.")
        ini = m.end()
        mnota = re.search(r"NOTA EXPLICATIVA", texto[ini:])
        fim = ini + mnota.start() if mnota else len(texto)
        corpo = re.sub(r"\s+", " ", texto[ini:fim]).strip()
        if not corpo:
            raise RuntimeError(f"RGI Regra {n} veio vazia na extração.")
        unidades.append(UnidadeNormativa(identificador=f"RGI Regra {n}", texto=corpo))
    return unidades


def _extrair_secao_rgi(pdf_bytes: bytes) -> str:
    """Concatena as páginas da seção RGI, detectadas dinamicamente (não por número fixo):
    começa na página com 'REGRA 1' + 'valor indicativo'; termina ao chegar na Seção I."""
    import pdfplumber

    paginas: list[str] = []
    capturando = False
    with pdfplumber.open(io.BytesIO(pdf_bytes)) as pdf:
        for idx, page in enumerate(pdf.pages):
            if idx > 60 and not capturando:
                raise RuntimeError("Início da seção RGI não encontrado nas primeiras páginas da NESH.")
            texto = page.extract_text() or ""
            up = texto.upper()
            if not capturando:
                if "REGRA 1" in up and "VALOR INDICATIVO" in up:
                    capturando = True
                    paginas.append(texto)
                continue
            if "ANIMAIS VIVOS E PRODUTOS DO REINO ANIMAL" in up and "REGRA" not in up:
                break  # fronteira: começou a Seção I
            paginas.append(texto)
            if len(paginas) > 20:
                raise RuntimeError("Fim da seção RGI não detectado — estrutura da NESH pode ter mudado.")
    return "\n".join(paginas)


# --- Soluções de Consulta / Divergência via SIJUT2 (Receita Federal) ---

SIJUT2_PERMALINK = "http://normas.receita.fazenda.gov.br/sijut2consulta/link.action?idAto={id_ato}"

_RE_LINHA = re.compile(r"<tr class='linhaResultados'>(.*?)</tr>", re.S)
_RE_ID_ATO = re.compile(r"link\.action\?(?:antigo=1&(?:amp;)?)?idAto=(\d+)")
_RE_TD = re.compile(r"<td[^>]*>(.*?)</td>", re.S)
_RE_TOTAL_PAGINAS = re.compile(r"de\s+(\d+)\s*<i[^>]*btnProximaPagina2", re.S)


def _limpar_celula(td_html: str) -> str:
    """Comentários fora, <br> vira quebra de linha, tags fora, entidades decodificadas."""
    sem_comentarios = re.sub(r"<!--.*?-->", "", td_html, flags=re.S)
    com_quebras = re.sub(r"<br\s*/?>", "\n", sem_comentarios, flags=re.I)
    sem_tags = re.sub(r"<[^>]+>", "", com_quebras)
    linhas = [ln.strip() for ln in html_mod.unescape(sem_tags).splitlines()]
    return "\n".join(ln for ln in linhas if ln)


def parse_sijut2_pagina(html: str) -> tuple[list[dict], int | None]:
    """Extrai os atos de uma página de listagem do SIJUT2.

    Retorna (atos, total_paginas). Cada ato: {id_ato, tipo, numero, orgao_emissor,
    data_publicacao, ementa}. Colunas da tabela: Tipo | Número | Órgão | Data | Ementa.
    """
    atos: list[dict] = []
    for bloco in _RE_LINHA.findall(html):
        # idAto vem ANTES de remover comentários: no HTML real ele só aparece dentro
        # dos <!-- <a href='link.action?idAto=N'> --> comentados.
        m_id = _RE_ID_ATO.search(bloco)
        # Comentários também contêm <td> duplicados — remover antes de parsear células,
        # senão as colunas deslocam.
        bloco_sem_comentarios = re.sub(r"<!--.*?-->", "", bloco, flags=re.S)
        celulas = [_limpar_celula(td) for td in _RE_TD.findall(bloco_sem_comentarios)]
        if m_id is None or len(celulas) < 5:
            continue  # linha de layout/estrutura inesperada — ignora, não inventa
        atos.append(
            {
                "id_ato": m_id.group(1),
                "tipo": celulas[0].replace("\n", " ").strip(),
                "numero": celulas[1].replace("\n", " ").strip(),
                "orgao_emissor": celulas[2].replace("\n", " ").strip(),
                "data_publicacao": celulas[3].replace("\n", " ").strip(),
                "ementa": celulas[4].strip(),
            }
        )
    m_total = _RE_TOTAL_PAGINAS.search(html)
    total = int(m_total.group(1)) if m_total else None
    return atos, total


def _ato_para_unidade(ato: dict) -> UnidadeNormativa:
    ano = ato["data_publicacao"][-4:] if len(ato["data_publicacao"]) >= 4 else "s/d"
    identificador = f"{ato['tipo']} {ato['orgao_emissor']} nº {ato['numero']}/{ano}"
    return UnidadeNormativa(
        identificador=identificador,
        texto=ato["ementa"],
        fonte_url=SIJUT2_PERMALINK.format(id_ato=ato["id_ato"]),
    )


@register("sijut2_sc")
def sijut2_sc_loader(fonte: FonteConfig) -> list[UnidadeNormativa]:
    """Coleta Soluções de Consulta/Divergência da listagem do SIJUT2 (RFB).

    A ementa completa (com o campo oficial 'Assunto:') vem inline na listagem — não é
    preciso abrir ato por ato. Paginação é GET puro via parâmetro p=N (~100 atos/página).

    params (config):
      filtro_assunto: só mantém atos cuja linha 'Assunto:' contenha este texto
                      (ex.: 'Classificação de Mercadorias'). Sem filtro = tudo.
      max_paginas:    teto de páginas a varrer (para smoke test). Sem teto = todas.
      delay_s:        pausa de cortesia entre páginas (default 1.0s).
    """
    params = fonte.params or {}
    filtro_assunto = (params.get("filtro_assunto") or "").strip().lower()
    max_paginas = params.get("max_paginas")
    delay_s = float(params.get("delay_s", 1.0))

    vistos: set[str] = set()
    # identificador -> unidade. A listagem vem em data de publicação DECRESCENTE, então a
    # primeira ocorrência de um identificador é a publicação mais recente — republicações/
    # retificações antigas do mesmo ato são descartadas para nunca gerar dois "vigentes"
    # com o mesmo identificador no mesmo lote (invariante do versionamento por vigência).
    por_identificador: dict[str, UnidadeNormativa] = {}
    total_paginas: int | None = None
    pagina = 1

    with httpx.Client(timeout=120, follow_redirects=True, headers=_UA) as client:
        while True:
            url = re.sub(r"([?&])p=\d+", rf"\g<1>p={pagina}", fonte.fonte_url)
            resp = client.get(url)
            resp.raise_for_status()
            atos, total = parse_sijut2_pagina(resp.text)
            if total_paginas is None and total is not None:
                total_paginas = total
            if not atos:
                break  # página vazia = fim (defensivo, mesmo se total não foi lido)

            for ato in atos:
                if ato["id_ato"] in vistos:
                    continue  # dedupe: listagem pode deslizar entre fetches
                vistos.add(ato["id_ato"])
                if filtro_assunto:
                    assunto = ato["ementa"].splitlines()[0].lower() if ato["ementa"] else ""
                    if filtro_assunto not in assunto:
                        continue
                unidade = _ato_para_unidade(ato)
                por_identificador.setdefault(unidade.identificador, unidade)

            limite = min(x for x in (total_paginas, max_paginas) if x is not None) \
                if (total_paginas or max_paginas) else None
            if pagina % 25 == 0:
                print(f"  [sijut2] página {pagina}/{limite or '?'} — {len(por_identificador)} atos no filtro até aqui", flush=True)
            if limite is not None and pagina >= limite:
                break
            pagina += 1
            time.sleep(delay_s)

    return list(por_identificador.values())


@register("rgi_nesh")
def rgi_nesh_loader(fonte: FonteConfig) -> list[UnidadeNormativa]:
    """Coleta as RGI a partir da NESH oficial (Receita Federal).

    fonte_url deve ser a PÁGINA da NESH (não o PDF direto) — o loader resolve o link do
    PDF vigente ali, baixa, isola a seção RGI e chunka nas 6 regras."""
    with httpx.Client(timeout=180, follow_redirects=True, headers=_UA) as client:
        pagina = client.get(fonte.fonte_url)
        pagina.raise_for_status()
        pdf_url = resolver_pdf_nesh(pagina.text, str(pagina.url))
        pdf_resp = client.get(pdf_url)
        pdf_resp.raise_for_status()
        pdf_bytes = pdf_resp.content
    secao = _extrair_secao_rgi(pdf_bytes)
    return parse_rgi_texto(secao)
