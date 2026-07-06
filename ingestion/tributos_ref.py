"""Carga da camada de Tributos (REFERÊNCIA) a partir da planilha do time (tax_calc.xlsx).

Abas: 'tax' (alíquotas por NCM), 'icms' (por UF), 'siscomex' (taxa por nº de adições).
Marcada como snapshot de referência datado — NÃO é a fonte oficial (TEC/TIPI). Idempotente
por data_referencia; grava linha de provenance em `normas` para citação.

Uso: mcp-server/.venv/Scripts/python.exe ingestion/tributos_ref.py <caminho.xlsx> [YYYY-MM-DD]
"""
from __future__ import annotations

import asyncio
import os
import pathlib
import re
import sys
from datetime import date

import asyncpg
import openpyxl
from dotenv import load_dotenv

ROOT = pathlib.Path(__file__).resolve().parent.parent
load_dotenv(ROOT / ".env")

ORIGEM = "tax_calc.xlsx (referência do time — não oficial)"


def _num(v) -> float | None:
    if v is None or str(v).strip() == "":
        return None
    try:
        return float(str(v).replace("%", "").replace(",", ".").strip())
    except ValueError:
        return None


def _ncm_fmt(v) -> str | None:
    d = re.sub(r"\D", "", str(v or ""))
    return f"{d[:4]}.{d[4:6]}.{d[6:8]}" if len(d) == 8 else None


def _linha(r: tuple, n: int) -> list:
    """Normaliza a linha para n colunas (read_only devolve tuplas curtas em linhas ragged)."""
    return list(r) + [None] * (n - len(r)) if len(r) < n else list(r)


def parse_planilha(caminho: pathlib.Path) -> dict:
    wb = openpyxl.load_workbook(caminho, read_only=True, data_only=True)
    tax, icms, sisc = [], [], []
    for row in wb["tax"].iter_rows(min_row=2, values_only=True):
        r = _linha(row, 12)
        ncm = _ncm_fmt(r[0])
        if not ncm:
            continue
        tax.append({"ncm": ncm, "ii": _num(r[4]), "ipi": _num(r[5]), "pis": _num(r[6]),
                    "cofins": _num(r[7]), "cide": (str(r[8]).strip() if r[8] else None),
                    "antidumping": (str(r[9]).strip() if r[9] else None),
                    "medidas": (str(r[10]).strip() if r[10] else None),
                    "trat": (str(r[11]).strip() if r[11] else None),
                    "descricao": (str(r[2]).strip() if r[2] else None)})
    for row in wb["icms"].iter_rows(min_row=2, values_only=True):
        r = _linha(row, 5)
        if r[0]:
            icms.append({"uf": str(r[0]).strip(), "estado": str(r[1]).strip() if r[1] else None,
                         "icms": _num(r[2]), "afrmm": _num(r[3]), "taxa_mm": _num(r[4])})
    for row in wb["siscomex"].iter_rows(min_row=2, values_only=True):
        r = _linha(row, 3)
        if r[0] is not None:
            sisc.append({"qtde": int(float(r[0])), "por_adicao": _num(r[1]), "total": _num(r[2])})
    wb.close()
    # dedup NCM (planilha pode repetir): mantém a 1ª ocorrência
    vistos, tax_u = set(), []
    for t in tax:
        if t["ncm"] not in vistos:
            vistos.add(t["ncm"]); tax_u.append(t)
    return {"tax": tax_u, "icms": icms, "siscomex": sisc}


async def carregar(caminho: pathlib.Path, ref: date) -> None:
    dados = parse_planilha(caminho)
    print(f"parse: {len(dados['tax'])} NCM, {len(dados['icms'])} UF, {len(dados['siscomex'])} faixas siscomex")
    conn = await asyncpg.connect(os.environ["DATABASE_URL"])
    try:
        if await conn.fetchval("SELECT 1 FROM tributos_ncm WHERE data_referencia=$1 LIMIT 1", ref):
            print(f"Referência {ref} já carregada — idempotente, nada a fazer.")
            return
        async with conn.transaction():
            await conn.executemany(
                "INSERT INTO tributos_ncm (ncm, ii, ipi, pis, cofins, cide, antidumping, "
                "medidas_compensatorias, tratamento_administrativo, descricao, origem, data_referencia) "
                "VALUES ($1,$2,$3,$4,$5,$6,$7,$8,$9,$10,$11,$12)",
                [(t["ncm"], t["ii"], t["ipi"], t["pis"], t["cofins"], t["cide"], t["antidumping"],
                  t["medidas"], t["trat"], t["descricao"], ORIGEM, ref) for t in dados["tax"]])
            await conn.executemany(
                "INSERT INTO icms_uf (uf, estado, icms, afrmm, taxa_utilizacao_mm, origem, data_referencia) "
                "VALUES ($1,$2,$3,$4,$5,$6,$7)",
                [(i["uf"], i["estado"], i["icms"], i["afrmm"], i["taxa_mm"], ORIGEM, ref) for i in dados["icms"]])
            await conn.executemany(
                "INSERT INTO taxa_siscomex (qtde_adicoes, valor_por_adicao, valor_total, origem, data_referencia) "
                "VALUES ($1,$2,$3,$4,$5) ON CONFLICT DO NOTHING",
                [(s["qtde"], s["por_adicao"], s["total"], ORIGEM, ref) for s in dados["siscomex"]])
            await conn.execute(
                "UPDATE normas SET data_vigencia_fim=$1 WHERE tipo_documento='tributos_ref' AND data_vigencia_fim IS NULL", ref)
            await conn.execute(
                "INSERT INTO normas (orgao, tipo_documento, identificador, texto, fonte_url, data_vigencia_inicio) "
                "VALUES ('RFB/CAMEX', 'tributos_ref', $1, $2, NULL, $3)",
                f"Tributos (referência) — snapshot {ref}",
                f"Alíquotas de importação (II/IPI/PIS/COFINS) por NCM + ICMS por UF + taxa Siscomex. "
                f"Fonte: planilha de referência do time ({len(dados['tax'])} NCM) — NÃO OFICIAL, pode estar "
                f"desatualizada. Fonte oficial de destino: TEC (CAMEX) + TIPI (RFB).", ref)
        print(f"Carga concluída (referência {ref}).")
    finally:
        await conn.close()


if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("Uso: python ingestion/tributos_ref.py <caminho.xlsx> [YYYY-MM-DD]")
        raise SystemExit(2)
    caminho = pathlib.Path(sys.argv[1])
    ref = date.fromisoformat(sys.argv[2]) if len(sys.argv) > 2 else date.today()
    asyncio.run(carregar(caminho, ref))
